import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook(r'D:/claude_project/eliaszeng.github.io/public/learning/12周学习计划_修订版.xlsx', read_only=True, data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'=== {sheet} ===')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 60:
            break
        vals = [str(c).strip() if c else '' for c in row]
        line = ' | '.join(v for v in vals if v)
        if line:
            print(line)
    print()
